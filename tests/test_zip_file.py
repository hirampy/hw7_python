import csv
import zipfile
from pypdf import PdfReader
from openpyxl import load_workbook
from tests.conftest import archive


def test_csv():
    with zipfile.ZipFile(archive, 'r') as zip_file:
        with zip_file.open('annual.csv') as csv_file:
            content = csv_file.read().decode('utf-8-sig')
            csvreader = list(csv.reader(content.splitlines()))
        assert csvreader[27][4] == 'Fixed tangible assets'


def test_pdf():
    with zipfile.ZipFile(archive, 'r') as zip_file:
        with zip_file.open('welcombonus.pdf') as pdf_file:
            reader = PdfReader(pdf_file)
            pdf_content = reader.pages[1].extract_text()
    assert 'бонус' in pdf_content


def test_xlsx():
    with zipfile.ZipFile(archive, 'r') as zip_file:
        with zip_file.open('file_XLSX_50.xlsx') as xlsx_file:
            workbook = load_workbook(xlsx_file)
            sheet = workbook.active
    assert sheet.cell(row=28, column=3).value == 'Vong'
